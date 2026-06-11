"""Pre-Flight-Checks (Pflicht vor jedem Crawl).

1. Katalog KONSUMIEREN, nicht entdecken: I14Y Public API (Behoerdenleistungen,
   ohne Auth) und eCH-0070-Leistungsinventar (XLSX) abrufen; Abdeckung pro
   kuratierter Leistung nach reports/coverage.md schreiben.
2. Rechtsflaeche pruefen: robots.txt jeder Quell-Domain gegen unseren
   User-Agent auswerten, Nutzungsbedingungs-Links dokumentieren; Funde nach
   reports/scraping-compliance.md. Bei Disallow wird die Leistung GESPERRT
   (der Crawler verweigert sie) und fuer Rueckfrage geflaggt.
"""
from __future__ import annotations

import io
import json
import time
import urllib.robotparser
from datetime import date
from urllib.parse import urlsplit

import httpx
import openpyxl

from .config import ROOT, ProcessSource, SourcesConfig

REPORTS = ROOT / "reports"
GATE_FILE = REPORTS / "raw" / "preflight-gate.json"

# UA-Token, auf das robots.txt-Regeln matchen (erste Komponente des User-Agent).
UA_TOKEN = "tessera"

# Bekannte Seiten mit rechtlichen Hinweisen / Nutzungsbedingungen je Domain.
# Diese werden im Compliance-Report verlinkt; die inhaltliche Pruefung bleibt
# eine dokumentierte Maintainer-Aufgabe (kein automatisches "alles ok").
KNOWN_TERMS = {
    "www.stadt-zuerich.ch": "https://www.stadt-zuerich.ch/de/impressum.html",
    "www.zh.ch": "https://www.zh.ch/de/impressum-rechtliches.html",
}


def _fetch_i14y(url: str, ua: str) -> list[str]:
    """Holt alle Behoerdenleistungs-Namen (de) aus der I14Y Public API."""
    names: list[str] = []
    page = 1
    with httpx.Client(headers={"User-Agent": ua}, timeout=30) as client:
        while True:
            r = client.get(url, params={"page": page, "pageSize": 100})
            r.raise_for_status()
            data = r.json().get("data", [])
            for svc in data:
                name = (svc.get("name") or {}).get("de") or ""
                kw = [
                    (k.get("de") or "")
                    for k in (svc.get("keywords") or [])
                    if isinstance(k, dict)
                ]
                names.append(" | ".join([name, *kw]).strip())
            total_pages = int(r.headers.get("x-paging-totalpages", page))
            if page >= total_pages:
                break
            page += 1
            time.sleep(0.5)
    return names


def _fetch_ech0070(url: str, ua: str) -> list[str]:
    """Holt die deutschen Leistungsbezeichnungen aus dem eCH-0070-Inventar."""
    with httpx.Client(headers={"User-Agent": ua}, timeout=60, follow_redirects=True) as client:
        r = client.get(url)
        r.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
    sheet = next((n for n in wb.sheetnames if "Leistungsinventar" in n), None)
    if sheet is None:
        raise RuntimeError(f"Kein Leistungsinventar-Sheet in {wb.sheetnames}")
    ws = wb[sheet]
    rows = ws.iter_rows(min_row=1, values_only=True)
    header = [str(c or "") for c in next(rows)]
    col = next(
        (i for i, h in enumerate(header) if h.startswith("Leistungsergebnis")), None
    )
    if col is None:
        raise RuntimeError(f"Spalte 'Leistungsergebnis' nicht gefunden: {header}")
    return [str(row[col]) for row in rows if row[col]]


def _matches(names: list[str], keywords: list[str]) -> list[str]:
    hits: list[str] = []
    for n in names:
        low = n.lower()
        if any(k.lower() in low for k in keywords):
            hits.append(n)
    return hits


def _robots_for_host(host: str, ua: str) -> tuple[urllib.robotparser.RobotFileParser | None, str]:
    """Laedt und parst robots.txt einer Domain. (parser, status_notiz)."""
    rp = urllib.robotparser.RobotFileParser()
    url = f"https://{host}/robots.txt"
    try:
        r = httpx.get(url, headers={"User-Agent": ua}, timeout=20)
    except httpx.HTTPError as exc:
        return None, f"robots.txt nicht abrufbar ({exc.__class__.__name__})"
    if r.status_code != 200:
        # Kein robots.txt => keine maschinenlesbare Einschraenkung.
        return None, f"robots.txt HTTP {r.status_code} (keine Regeln)"
    rp.parse(r.text.splitlines())
    return rp, "robots.txt geladen"


def run_preflight(cfg: SourcesConfig, only: list[str] | None = None) -> dict[str, dict]:
    """Fuehrt beide Checks aus, schreibt Reports und das Gate-File.

    Rueckgabe: {proc_id: {"allowed": bool, "blocked_urls": [...]}}
    """
    today = date.today().isoformat()
    ua = cfg.crawler.user_agent
    procs = [p for p in cfg.processes if not only or p.id in only]

    # --- 1. Katalog-Abdeckung -------------------------------------------------
    i14y_names: list[str] = []
    i14y_err = ""
    try:
        i14y_names = _fetch_i14y(cfg.catalog.i14y_publicservices_url, ua)
    except Exception as exc:  # Report statt Crash: Befund gehoert in coverage.md
        i14y_err = f"{exc.__class__.__name__}: {exc}"
    ech_names: list[str] = []
    ech_err = ""
    try:
        ech_names = _fetch_ech0070(cfg.catalog.ech0070_inventory_url, ua)
    except Exception as exc:
        ech_err = f"{exc.__class__.__name__}: {exc}"

    lines = [
        "# Abdeckung der kuratierten Leistungen in den Katalogen",
        "",
        f"Stand: {today} — erzeugt durch `tessera preflight` (Katalog wird",
        "konsumiert, nicht entdeckt; kuratierte Liste: `sources.yaml`).",
        "",
        f"- I14Y Public API (Behoerdenleistungen): "
        + (f"{len(i14y_names)} Eintraege" if not i14y_err else f"FEHLER — {i14y_err}"),
        f"- eCH-0070-Leistungsinventar (V4.2.0, XLSX): "
        + (f"{len(ech_names)} Eintraege" if not ech_err else f"FEHLER — {ech_err}"),
        "",
        "| Leistung (kuratiert) | I14Y-Treffer | eCH-0070-Treffer |",
        "|---|---|---|",
    ]
    for p in procs:
        i_hits = _matches(i14y_names, p.catalog_keywords)
        e_hits = _matches(ech_names, p.catalog_keywords)
        fmt = lambda hits: "; ".join(h[:60] for h in hits[:4]) + (
            f" (+{len(hits) - 4})" if len(hits) > 4 else ""
        ) if hits else "—  (kommunale Leistung, kein Bundes-Eintrag)"
        lines.append(f"| `{p.id}` ({p.service_name}) | {fmt(i_hits)} | {fmt(e_hits)} |")
    lines += [
        "",
        "Hinweis: Beide Kataloge sind auf Bundes-/Kantonsebene gepflegt; rein",
        "kommunale Leistungen der Stadt Zuerich koennen ohne Treffer bleiben.",
        "Das ist ein dokumentierter Befund, kein Fehler.",
        "",
    ]
    (REPORTS / "coverage.md").write_text("\n".join(lines), encoding="utf-8")

    # --- 2. Rechtsflaeche (robots.txt + ToU) ----------------------------------
    gate: dict[str, dict] = {}
    hosts = sorted({urlsplit(u).netloc for p in procs for u in p.official_urls})
    robots: dict[str, tuple[urllib.robotparser.RobotFileParser | None, str]] = {
        h: _robots_for_host(h, ua) for h in hosts
    }

    clines = [
        "# Scraping-Compliance (robots.txt & Nutzungsbedingungen)",
        "",
        f"Stand: {today} — erzeugt durch `tessera preflight`.",
        f"User-Agent: `{ua}`",
        f"Rate-Limit: {cfg.crawler.delay_seconds}s Pause zwischen Requests;",
        "keine Umgehung technischer Schutzmassnahmen; nur oeffentliche Seiten,",
        "keine personenbezogenen Daten in URLs oder Logs.",
        "",
        "## Domains",
        "",
        "| Domain | robots.txt | Nutzungsbedingungen |",
        "|---|---|---|",
    ]
    for h in hosts:
        _, note = robots[h]
        terms = KNOWN_TERMS.get(h, "—")
        terms_md = f"[{terms}]({terms}) — manuelle Pruefung Maintainer" if terms != "—" else "— (Link nachtragen)"
        clines.append(f"| {h} | {note} | {terms_md} |")

    clines += ["", "## Geprüfte URLs", "", "| Leistung | URL | robots-Verdikt |", "|---|---|---|"]
    for p in procs:
        blocked: list[str] = []
        for u in p.official_urls:
            rp, _ = robots[urlsplit(u).netloc]
            allowed = rp.can_fetch(UA_TOKEN, u) if rp else True
            if not allowed:
                blocked.append(u)
            clines.append(f"| `{p.id}` | {u} | {'erlaubt' if allowed else '**DISALLOW — nicht crawlen**'} |")
        gate[p.id] = {"allowed": not blocked, "blocked_urls": blocked, "checked_at": today}
    clines += [
        "",
        "Verdikt-Logik: Eine Leistung wird nur gecrawlt, wenn ALLE ihre URLs",
        "fuer unseren User-Agent erlaubt sind. Bei Disallow: Leistung gesperrt,",
        "Flag im Report — Ruecksprache mit dem Maintainer noetig.",
        "",
    ]
    (REPORTS / "scraping-compliance.md").write_text("\n".join(clines), encoding="utf-8")

    GATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    GATE_FILE.write_text(json.dumps(gate, indent=2, ensure_ascii=False), encoding="utf-8")
    return gate


def load_gate() -> dict[str, dict]:
    if not GATE_FILE.exists():
        return {}
    return json.loads(GATE_FILE.read_text(encoding="utf-8"))


def require_allowed(proc: ProcessSource) -> None:
    """Crawl-Gate: ohne frischen, positiven Preflight wird nicht gecrawlt."""
    gate = load_gate()
    entry = gate.get(proc.id)
    if entry is None:
        raise SystemExit(
            f"[{proc.id}] Kein Preflight-Ergebnis gefunden — zuerst `tessera preflight` ausfuehren."
        )
    if not entry["allowed"]:
        raise SystemExit(
            f"[{proc.id}] robots.txt verbietet das Crawlen von {entry['blocked_urls']} — "
            "Leistung gesperrt; bitte Maintainer fragen (siehe reports/scraping-compliance.md)."
        )
